import streamlit as st
import pandas as pd
from faker import Faker
import re
import io
import phonenumbers
from phonenumbers import geocoder
from openpyxl.styles import PatternFill, Font, Alignment

# Initialize Faker with US locale
fake = Faker('en_US')

# Constants
DEFAULT_PHONE = "(212) 555-0100"
DEFAULT_COUNT = 100
DEFAULT_FILENAME = "generated_profiles.xlsx"
MAX_RECORDS = 10000

US_STATES = {
    'AK': 'Alaska', 'AL': 'Alabama', 'AR': 'Arkansas', 'AZ': 'Arizona', 'CA': 'California', 'CO': 'Colorado',
    'CT': 'Connecticut', 'DC': 'District of Columbia', 'DE': 'Delaware', 'FL': 'Florida', 'GA': 'Georgia',
    'HI': 'Hawaii', 'IA': 'Iowa', 'ID': 'Idaho', 'IL': 'Illinois', 'IN': 'Indiana', 'KS': 'Kansas', 'KY': 'Kentucky',
    'LA': 'Louisiana', 'MA': 'Massachusetts', 'MD': 'Maryland', 'ME': 'Maine', 'MI': 'Michigan', 'MN': 'Minnesota',
    'MO': 'Missouri', 'MS': 'Mississippi', 'MT': 'Montana', 'NC': 'North Carolina', 'ND': 'North Dakota',
    'NE': 'Nebraska', 'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NV': 'Nevada',
    'NY': 'New York', 'OH': 'Ohio', 'OK': 'Oklahoma', 'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island',
    'SC': 'South Carolina', 'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah', 'VA': 'Virginia',
    'VT': 'Vermont', 'WA': 'Washington', 'WI': 'Wisconsin', 'WV': 'West Virginia', 'WY': 'Wyoming'
}

REVERSE_US_STATES = {v: k for k, v in US_STATES.items()}

AREA_CODES_BY_STATE = {
    "NJ": ["201", "551", "609", "640", "732", "848", "856", "862", "908", "973"], "DC": ["202", "771"], 
    "CT": ["203", "475", "860", "959"], "AL": ["205", "251", "256", "334", "659", "938"], "ME": ["207"], 
    "ID": ["208", "986"], "CA": ["209", "213", "279", "310", "323", "341", "350", "369", "408", "415", "424", "442", "510", "530", "559", "562", "619", "626", "628", "650", "657", "661", "669", "707", "714", "738", "747", "760", "805", "818", "820", "831", "840", "858", "909", "916", "925", "949", "951"], 
    "TX": ["210", "214", "254", "281", "325", "346", "361", "409", "430", "432", "469", "512", "682", "713", "726", "737", "806", "817", "830", "832", "903", "915", "936", "940", "945", "956", "972", "979"], 
    "NY": ["212", "315", "329", "332", "347", "363", "516", "518", "585", "607", "631", "646", "680", "716", "718", "838", "845", "914", "917", "929", "934"], 
    "PA": ["215", "223", "267", "272", "412", "445", "484", "570", "582", "610", "717", "724", "814", "835", "878"], 
    "OH": ["216", "220", "234", "283", "326", "330", "380", "419", "440", "513", "567", "614", "740", "937"], 
    "IL": ["217", "224", "309", "312", "331", "447", "464", "618", "630", "708", "730", "773", "779", "815", "847", "872"], 
    "MN": ["218", "320", "507", "612", "651", "763", "952"], "IN": ["219", "260", "317", "463", "574", "765", "812", "930"], 
    "LA": ["225", "318", "337", "504", "985"], "MD": ["227", "240", "301", "410", "443", "667"], "MS": ["228", "601", "662", "769"], 
    "GA": ["229", "404", "470", "478", "678", "706", "762", "770", "912", "943"], "MI": ["231", "248", "269", "313", "517", "586", "616", "734", "810", "906", "947", "989"], 
    "MO": ["235", "314", "417", "557", "573", "636", "660", "816", "975"], 
    "FL": ["239", "305", "321", "324", "352", "386", "407", "448", "561", "645", "656", "689", "727", "728", "754", "772", "786", "813", "850", "863", "904", "941", "954"], 
    "NC": ["252", "336", "704", "743", "828", "910", "919", "980", "984"], "WI": ["262", "274", "353", "414", "534", "608", "715", "920"], 
    "KY": ["270", "364", "502", "606", "859"], "VA": ["276", "434", "540", "571", "686", "703", "757", "804", "826", "948"], 
    "DE": ["302"], "CO": ["303", "719", "720", "748", "970", "983"], "WV": ["304", "681"], "WY": ["307"], 
    "NE": ["308", "402", "531"], "KS": ["316", "620", "785", "913"], "IA": ["319", "515", "563", "641", "712"], 
    "AR": ["327", "479", "501", "870"], "MA": ["339", "351", "413", "508", "617", "774", "781", "857", "978"], 
    "UT": ["385", "435"], "RI": ["401"], "OK": ["405", "539", "572", "580", "918"], "MT": ["406"], 
    "TN": ["423", "615", "629", "731", "865", "901", "931"], "OR": ["458", "503", "541", "971"], 
    "AZ": ["480", "520", "602", "623", "928"], "NM": ["505", "575"], "NH": ["603"], "SD": ["605"], 
    "MP": ["670"], "GU": ["671"], "ND": ["701"], "NV": ["702", "725", "775"], "PR": ["787", "939"], "VT": ["802"], 
    "SC": ["803", "821", "839", "843", "854", "864"], "HI": ["808"], "AK": ["907"], 
    "WA": ["206", "253", "360", "425", "509", "564"]
}

def extract_state_from_geocoder(geo_str: str) -> str:
    if not geo_str: return ""
    if geo_str in REVERSE_US_STATES: return REVERSE_US_STATES[geo_str]
    parts = geo_str.split(', ')
    if len(parts) == 2:
        abbr = parts[1].strip().upper()
        if abbr in US_STATES: return abbr
    return ""

st.set_page_config(page_title="USA Profile & Phone Number Generator", layout="wide", page_icon="🇺🇸")

# Inject Custom CSS
st.markdown("""
    <style>
    /* Premium aesthetics */
    body {
        font-family: 'Inter', system-ui, sans-serif;
    }
    
    /* Header styling */
    .premium-header {
        text-align: center;
        padding-top: 1rem;
        padding-bottom: 2rem;
    }
    .premium-header h1 {
        font-weight: 700;
        color: #F0F2F6;
        margin-bottom: 0.2rem;
    }
    .premium-header p {
        color: #9BA1A6;
        font-size: 1.1rem;
    }

    /* Refined Metric styling */
    div[data-testid="stMetricValue"] {
        color: #00FFC2;
    }
    
    /* Footer */
    .footer {
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
        background-color: #1E232E;
        color: #9BA1A6;
        text-align: center;
        padding: 10px;
        font-size: 0.8rem;
        z-index: 1000;
        border-top: 1px solid #333;
    }
    
    /* Hide some default streamlit stuff */
    footer {visibility: hidden;}
    
    </style>
""", unsafe_allow_html=True)

def parse_phone_number(phone_str: str) -> str:
    """Extracts exactly 10 digits from the string, ignoring non-digit characters. Returns empty string if invalid."""
    digits = re.sub(r'\D', '', phone_str)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return digits
    return ""

def is_valid_us_number(phone_str: str) -> bool:
    """Checks if a phone number format is valid according to NANPA rules (offline library check)."""
    try:
        parsed = phonenumbers.parse(phone_str, "US")
        return phonenumbers.is_valid_number(parsed)
    except phonenumbers.NumberParseException:
        return False

def format_phone(digits: str) -> str:
    """Formats a 10-digit phone number integer into (XXX) XXX-XXXX."""
    return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"

def generate_profiles(base_phone: str, count: int, state_filter: str, selected_cols: list) -> pd.DataFrame:
    """Generates a dataframe of synthetic profiles."""
    records = []
    base_int = int(base_phone)
    
    for i in range(count):
        current_phone = base_int + i
        raw_phone = str(current_phone).zfill(10)
        formatted_phone = format_phone(raw_phone)
        
        # Valid format check
        # We prepend '+1' so phonenumbers parses it easily, but just standard format is fine.
        is_valid = is_valid_us_number(f"+1{raw_phone}")
        
        # Faker state loop if filtered
        state_abbr = fake.state_abbr()
        if state_filter and state_filter.upper() != "ALL":
            # Just keep generating until we hit the state, or use a specific locale. 
            # Faker is random, so generating until match is slow.
            # Instead, we just assign it to the selected state!
            state_abbr = state_filter.upper()
            
        zip_code = fake.zipcode()
        if state_filter and state_filter.upper() != "ALL" and hasattr(fake, 'zipcode_in_state'):
            try:
                zip_code = fake.zipcode_in_state(state_abbr)
            except:
                pass
                
        record = {
            "Phone Number (Raw)": raw_phone,
            "Phone Number (Formatted)": formatted_phone,
            "Format Valid": "✅" if is_valid else "❌",
            "First Name": fake.first_name(),
            "Last Name": fake.last_name(),
            "Street Address": fake.street_address(),
            "City": fake.city(), # Note: City remains nationally random as Faker lacks city_in_state out of box
            "State": state_abbr,
            "ZIP Code": zip_code
        }
        
        # Filter by selected columns
        filtered_record = {k: v for k, v in record.items() if k in selected_cols}
        records.append(filtered_record)
        
    return pd.DataFrame(records)

def create_styled_excel(df: pd.DataFrame) -> bytes:
    """Creates a styled Excel file from the dataframe and returns it as bytes."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Profiles')
        workbook = writer.book
        worksheet = writer.sheets['Profiles']
        
        header_font = Font(name='Inter', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0E1117', end_color='0E1117', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        
        for col_idx, column_cells in enumerate(worksheet.iter_cols(min_row=1, max_row=1, max_col=worksheet.max_column), 1):
            for cell in column_cells:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                
        worksheet.freeze_panes = "A2"
        for col_idx, col in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max(max_length + 2, 12)
    return output.getvalue()

def main():
    st.markdown('''
        <div class="premium-header">
            <h1>🇺🇸 USA Profile & Phone Number Generator</h1>
            <p>Generate synthetic contact datasets with structural validation.</p>
        </div>
    ''', unsafe_allow_html=True)
    
    if "start_phone_input" not in st.session_state:
        st.session_state.start_phone_input = DEFAULT_PHONE
        
    # Input Card container
    with st.container():
        st.subheader("Data Configuration")
        
        current_state_sel = st.session_state.get("state_filter", "Auto-Detect from Phone")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            start_phone_input = st.text_input("Starting Phone Number", key="start_phone_input", help="Accepts raw digits or formatted (e.g. 2125550100)")
            parsed_phone = parse_phone_number(start_phone_input)
            
            # Inline validation feedback
            detected_state = ""
            is_valid = False
            mismatch = False
            
            if parsed_phone:
                is_valid = is_valid_us_number(f"+1{parsed_phone}")
                if is_valid:
                    parsed_obj = phonenumbers.parse(f"+1{parsed_phone}", "US")
                    geo_str = geocoder.description_for_number(parsed_obj, "en")
                    detected_state = extract_state_from_geocoder(geo_str)
                    
                    if current_state_sel != "Auto-Detect from Phone" and current_state_sel != "ALL (Random States)":
                        sel_abbr = current_state_sel.split(" ")[0]
                        if detected_state and detected_state != sel_abbr:
                            mismatch = True
                            st.warning(f"⚠️ Area code {parsed_phone[:3]} belongs to {US_STATES.get(detected_state, detected_state)}, not {US_STATES.get(sel_abbr, sel_abbr)}. Choose a matching area code below or change the target state.")
                    
                    if not mismatch:
                        if detected_state:
                            st.success(f"✅ Valid format. Detected Area: **{US_STATES.get(detected_state, detected_state)}**")
                        else:
                            st.success("✅ Valid 10-digit US number format.")
                else:
                    st.warning("⚠️ 10 digits found, but structural format (area code/exchange) is invalid.")
            else:
                st.error("❌ Must contain exactly 10 digits.")
                
            resolved_state = detected_state if current_state_sel == "Auto-Detect from Phone" else (
                "ALL" if current_state_sel == "ALL (Random States)" else current_state_sel.split(" ")[0]
            )
            
            if resolved_state != "ALL" and resolved_state in AREA_CODES_BY_STATE:
                state_acs = AREA_CODES_BY_STATE[resolved_state]
                display_acs = state_acs[:10]
                more = f" (+{len(state_acs)-10} more)" if len(state_acs) > 10 else ""
                st.caption(f"Valid area codes for {US_STATES.get(resolved_state, resolved_state)}: {', '.join(display_acs)}{more}")
                
                def apply_area_code():
                    ac = st.session_state.ac_selector
                    if ac:
                        current = parse_phone_number(st.session_state.start_phone_input)
                        rest = current[3:] if len(current) == 10 else "5550100"
                        st.session_state.start_phone_input = f"{ac}{rest}"

                st.selectbox("Use a valid area code:", options=[""] + state_acs, key="ac_selector", on_change=apply_area_code, label_visibility="collapsed")
        
        with col2:
            state_options = ["Auto-Detect from Phone"] + [f"{k} — {v}" for k, v in US_STATES.items()] + ["ALL (Random States)"]
            state_filter_selection = st.selectbox("Target Generation State", options=state_options, index=0, key="state_filter", help="By default, uses the state of the area code you entered.")

        with col3:
            count_input = st.number_input("Number of Profiles", min_value=1, max_value=MAX_RECORDS, value=DEFAULT_COUNT, step=100)
            
        with col4:
            filename_input = st.text_input("Output File Name", value=DEFAULT_FILENAME)
            
    # Advanced Options
    with st.expander("⚙️ Advanced Options"):
        all_cols = ["Phone Number (Raw)", "Phone Number (Formatted)", "Format Valid", "First Name", "Last Name", "Street Address", "City", "State", "ZIP Code"]
        selected_cols = st.multiselect("Select Columns to Generate", options=all_cols, default=all_cols)

    st.markdown("<br>", unsafe_allow_html=True)
    
    # Action Row
    generate_disabled = not bool(parsed_phone) or not selected_cols or mismatch
    
    col_act1, col_act2 = st.columns([2, 10])
    with col_act1:
        generate_btn = st.button("🚀 Generate Profiles", type="primary", disabled=generate_disabled, use_container_width=True)

    # Results Area
    if generate_btn:
        if mismatch:
            st.error("Cannot generate: area code and target state do not match.")
        else:
            with st.status("Generating synthetic data...", expanded=True) as status:
                st.write("Initializing generation loop...")
                df = generate_profiles(parsed_phone, count_input, resolved_state, selected_cols)
                st.write("Validating structural format of numbers...")
                st.write("Compiling DataFrame...")
                status.update(label="Data generation complete!", state="complete", expanded=False)
            
            # Save to session state to persist through filter interactions
            st.session_state['generated_df'] = df
            st.session_state['current_parsed_phone'] = parsed_phone
            st.toast("✅ Generation Successful!", icon="🎉")

    if 'generated_df' in st.session_state:
        df = st.session_state['generated_df']
        original_parsed_phone = st.session_state.get('current_parsed_phone', parsed_phone)

        # State Filter UI
        st.markdown("---")
        
        if "State" in df.columns:
            available_states = sorted(df["State"].dropna().unique())
            
            def format_state(abbr):
                return f"{abbr} — {US_STATES.get(abbr, abbr)}"
                
            selected_states = st.multiselect(
                "Filter by State",
                options=available_states,
                format_func=format_state,
                help="Select one or more states to filter the generated results."
            )
            
            if selected_states:
                filtered_df = df[df["State"].isin(selected_states)]
            else:
                filtered_df = df
        else:
            filtered_df = df
            selected_states = []

        if len(filtered_df) == 0:
            st.info("No profiles match the selected states.")
        else:
            st.subheader("Summary Metrics")
            valid_count = len(filtered_df[filtered_df["Format Valid"] == "✅"]) if "Format Valid" in filtered_df.columns else "N/A"
            
            # Recalculate ranges for filtered view based on available phones
            if "Phone Number (Raw)" in filtered_df.columns:
                phone_ints = filtered_df["Phone Number (Raw)"].astype(int)
                min_phone = format_phone(str(phone_ints.min()).zfill(10))
                max_phone = format_phone(str(phone_ints.max()).zfill(10))
            else:
                min_phone = "N/A"
                max_phone = "N/A"
            
            m_col1, m_col2, m_col3, m_col4 = st.columns(4)
            m_col1.metric("Total Records", f"{len(filtered_df):,}")
            m_col2.metric("Start Range", min_phone)
            m_col3.metric("End Range", max_phone)
            m_col4.metric("Format Valid Numbers", f"{valid_count:,}" if valid_count != "N/A" else "N/A")
            
            if selected_states:
                st.caption(f"Showing {len(filtered_df):,} of {len(df):,} total generated records.")
            
            st.subheader("Data Preview")
            st.data_editor(filtered_df, hide_index=True, use_container_width=True)
            
            # Export Buttons
            if not filename_input.endswith('.xlsx'):
                filename_input += '.xlsx'
                
            dl_col1, dl_col2 = st.columns(2)
            
            with dl_col1:
                excel_bytes_all = create_styled_excel(df)
                st.download_button(
                    label="📥 Download All (Original)",
                    data=excel_bytes_all,
                    file_name=filename_input,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_all",
                    use_container_width=True
                )
                
            with dl_col2:
                if selected_states:
                    suffix = "_".join(selected_states[:3])
                    if len(selected_states) > 3:
                        suffix = "filtered"
                    filtered_filename = filename_input.replace('.xlsx', f'_{suffix}.xlsx')
                else:
                    filtered_filename = filename_input
                    
                excel_bytes_filtered = create_styled_excel(filtered_df)
                st.download_button(
                    label="📥 Download Filtered" if selected_states else "📥 Download Current View",
                    data=excel_bytes_filtered,
                    file_name=filtered_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_filtered",
                    type="primary",
                    use_container_width=True
                )
        
    st.markdown('<div class="footer">Version 2.0 | USA Profile Generator (Fictional Data Only)</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    main()
